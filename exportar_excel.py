import pandas as pd
import sqlite3
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Ruta de la base de datos
db_path = "instance/drogas.db"

# Conexión a SQLite
conn = sqlite3.connect(db_path)

# Leer la tabla 'droga' completa
df = pd.read_sql_query("SELECT * FROM droga", conn)

# Exportar a Excel
excel_path = "drogas_exportadas.xlsx"
df.to_excel(excel_path, index=False)

# Agregar filtros al Excel usando openpyxl
wb = load_workbook(excel_path)
ws = wb.active

# Determinar rango de la tabla (por ejemplo, A1:H{última fila})
num_rows = ws.max_row
num_cols = ws.max_column
last_col_letter = chr(64 + num_cols)

table = Table(displayName="TablaDrogas", ref=f"A1:{last_col_letter}{num_rows}")

# Estilo de tabla
style = TableStyleInfo(
    name="TableStyleMedium9",
    showRowStripes=True,
    showColumnStripes=False
)
table.tableStyleInfo = style

# Agregar la tabla al sheet
ws.add_table(table)

# Guardar Excel con filtros
wb.save(excel_path)
wb.close()

print(f"✅ Exportación completa: {excel_path}")
